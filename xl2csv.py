import os
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import win32com.client
import csv
import statistics as stats


class Excel:
    # 엑셀 파일 불러오기
    def load_file(self, filename):
        """
        :param filename: (str) 불러올 엑셀 파일 이름
        :return: <Workbook> 엑셀 파일의 워크북
        """
        wb = Workbook()
        if os.path.exists(filename):
            wb = load_workbook(filename, data_only=True)
        return wb

    # 워크시트 불러오기
    def load_worksheet(self, wb, sheetname):
        """
        :param wb: (Workbook) 워크시트를 부를 워크북
        :param sheetname: (str) 생성할 워크시트 이름
        :return: <Worksheet> 생성한 워크시트
        """
        if sheetname in wb.sheetnames:
            ws = wb[sheetname]
        elif sheetname == 'none':
            ws = wb.active
        else:
            ws = wb.create_sheet(sheetname)
        return ws

    # 데이터 읽어오기
    def get_value(self, ws, v, cell_range):
        """
        :param ws: (Worksheet) 데이터를 읽을 워크시트
        :param v: (bool) 데이터 읽을 방향(true이면 수직방향)
        :param cell_range: (str) 읽을 셀 범위
        :return: (list) 읽은 데이터의 리스트
        """
        val = []

        if cell_range is None:
            if v:
                for col in ws.columns:
                    tmp = []
                    for cell in col:
                        tmp.append(cell.value)
                    val.append(tmp)
            else:
                for row in ws.rows:
                    tmp = []
                    for cell in row:
                        tmp.append(cell.value)
                    val.append(tmp)
        # 단일 셀에 입력하는 경우
        elif ':' not in ''.join(map(str, cell_range)):
            val.append(ws[''.join(map(str, cell_range.split(':')))])
        else:
            for r in range(len(cell_range)):
                cell_range_tmp = ''.join(map(str, cell_range[r]))
                head_range = cell_range_tmp.split(':')[0]
                tail_range = cell_range_tmp.split(':')[1]
                min_col = column_index_from_string(ws[head_range].column)
                min_row = ws[head_range].row
                max_col = column_index_from_string(ws[tail_range].column)
                max_row = ws[tail_range].row

                if v:
                    for row in ws.iter_cols(min_row=min_row, min_col=min_col,
                                            max_row=max_row, max_col=max_col):
                        tmp = []
                        for cell in row:
                            try:
                                tmp.append(''.join(map(str, cell.value)))
                            except TypeError:
                                pass
                        val.append(tmp)
                else:
                    for col in ws.iter_rows(min_row=min_row, min_col=min_col,
                                            max_row=max_row, max_col=max_col):
                        tmp = []
                        for cell in col:
                            try:
                                tmp.append(''.join(map(str, cell.value)))
                            except TypeError:
                                pass
                        val.append(tmp)
        return val

    # CSV로 변환
    def to_csv(self, data: list, filename):
        """
        :param data: (list) csv로 변환할 값
        :param filename: (string) 생성할 csv 파일 이름
        :return: True
        """
        filename = ''.join(map(str, filename))
        if 'csv' in filename:
            fn = filename
        elif 'xlsx' in filename:
            fn = ''.join(map(str, filename.split('.')[0]))+'.csv'

        f = open(fn, 'w', encoding='utf-8', newline='')
        wr = csv.writer(f)
        for cell in data:
            if cell is None:
                cell = ''
            wr.writerow(cell)
        f.close()

        return True

    # write mode
    # 엑셀 파일 로드-> 기존 파일 or 파일 생성 판별할 수 있어야함: load_file
    # 시트 로드 -> 기존 시트 or 시트 생성 판별할 수 있어야함: load_worksheet
    # 데이터 쓰기 & 엑셀 저장 -> 방향에 따라 구현하기
    def set_value(self, wb, ws, cell_range, filename, variable, v):
        """
        :param wb: (Workbook) 쓸 워크북 이름
        :param ws: (Worksheet) 쓸 워크시트 이름
        :param cell_range: (str) 쓸 셀의 범위
        :param filename: (str) 쓸 엑셀파일 이름
        :param variable: (list) 쓸 값
        :param v: (bool) 읽는 방향(true이면 수직방향)
        :return: (Workbook)
        """
        if cell_range is None:
            k = 0
            for row in ws.rows:
                for cell in row:
                    try:
                        cell.value = variable[k]
                        k += 1
                    except IndexError:
                        pass

        elif ':' not in ''.join(map(str, cell_range)):
            ws[''.join(map(str, cell_range.split(':')))] = variable[0]
        else:
            k = 0
            for r in range(len(cell_range)):
                cell_range_tmp = ''.join(map(str, cell_range[r]))
                head_range = cell_range_tmp.split(':')[0]
                tail_range = cell_range_tmp.split(':')[1]
                min_col = column_index_from_string(ws[head_range].column)
                min_row = ws[head_range].row
                max_col = column_index_from_string(ws[tail_range].column)
                max_row = ws[tail_range].row

                if v:
                    for i in range(min_col, max_col + 1):
                        for j in range(min_row, max_row + 1):
                            try:
                                ws.cell(row=j, column=i).value = variable[k]
                            except IndexError:
                                ws.cell(row=j, column=i).value = None
                            k += 1
                else:
                    for i in range(min_row, max_row + 1):
                        for j in range(min_col, max_col + 1):
                            try:
                                ws.cell(row=i, column=j).value = variable[k]
                            except IndexError:
                                ws.cell(row=i, column=j).value = None
                            k += 1
        wb.save(filename)
        return wb

################################################################################
    # function mode
    # csv file 열기
    def load_csv_file(self, filename):
        """
        :param filename: (str) 가져올 csv 파일 이름
        :return: (file) csv 파일
        """
        f = open(filename, 'r', encoding='utf-8')
        return f

    # csv 데이터 읽어오기
    def get_csv_value(self, file):
        """
        :param file: (file) 데이터를 읽을 csv 파일 이름
        :return: (list) csv 파일 내용
        """
        val = []

        for row in csv.reader(file):
            val.append(row)
        return val

    # excel로 변환 + 연산(function)
    def do_function(self, data: list, ftype, cell_range):
        """
        :param data: (list) csv 파일에 있던 값
        :param ftype: (str) 값에 취할 연산 종류
        :param cell_range: (str) 연산할 셀의 범위
        :return: True
        """
        wb = Workbook(data_only=True)
        ws = wb.active

        print(data)
        # 받은 list가 str형이기 때문에 int(float)형으로 변환
        for i, val_i in enumerate(data):
            for j, val_j in enumerate(val_i):
                data[i][j] = float(data[i][j])

        for row in data:
            ws.append(row)

        val = []
        cell_ranges = ''.join(map(str, cell_range)).split(':')
        cells = ws[cell_ranges[0]:cell_ranges[1]]

        if cell_ranges[0][0] == cell_ranges[1][0]:
            next_cell = cell_ranges[1][0] + str(int(cell_ranges[1][1])+1)
        else:
            next_cell = get_column_letter(column_index_from_string
                                          (cell_ranges[1][0])+1) \
                        + cell_ranges[1][1]

        # 연산할 값들을 val에 저장
        for row in cells:
            for cell in row:
                val.append(float(cell.value))

        # 연산 수행
        if ftype == 'SUM':
            # ws[next_cell] = sum(val)
            ws[next_cell] = '=SUM({}:{})'.format(cell_ranges[0], cell_ranges[1])
        elif ftype == 'MIN':
            ws[next_cell] = min(val)
        elif ftype == 'MAX':
            ws[next_cell] = max(val)
        elif ftype == 'AVERAGE':
            ws[next_cell] = stats.mean(val)
        elif ftype == 'MEDIAN':
            ws[next_cell] = stats.median(val)

        val = []
        for row in ws.rows:
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            val.append(tmp)

        return val

################################################################################
    def do_macro(self, filename, macroname):
        """
        :param filename: (str) 매크로를 실행할 엑셀파일 이름
        :param macroname: (str) 실행할 매크로 이름
        :return: none
        """
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Workbooks.Open(Filename=filename, ReadOnly=1)
        xl.Application.Run(macroname)
        xl.Workbooks(1).Close(SaveChanges=1)
        xl.Application.Quit()
        del xl


################################################################################
def main():
    global lst
    parser = argparse.ArgumentParser()
    parser.add_argument('operation', type=str, help='read/write',
                        choices=['read', 'write'])
    parser.add_argument('filename', type=str, help='file name')
    parser.add_argument('--sheetname', type=str, help='sheet name',
                        default='none')
    parser.add_argument('--range', type=str, help='cell range',
                        action='append')
    parser.add_argument('-v', '--vertical', help='access vertically',
                        action='store_true')
    parser.add_argument('--type', type=str, help
                        ='types of function(SUM, MIN, MAX, AVERAGE, MEDIAN)',
                        choices=['SUM', 'MIN', 'MAX', 'AVERAGE', 'MEDIAN'],
                        default=None)
    parser.add_argument('--macroname', type=str, help='macro name')
    parser.add_argument('--input', type=list, nargs='+', help='input value')
    args = parser.parse_args()

    if args.input is not None:
        lst = []
        for i in range(len(args.input)):
            lst.append(''.join(map(str, args.input[i])))

    xl = Excel()

    if args.operation == 'read':
        wb = xl.load_file(args.filename)
        ws = xl.load_worksheet(wb=wb, sheetname=args.sheetname)
        data = xl.get_value(ws, args.vertical, args.range)
        xl.to_csv(data=data, filename=args.filename)
    elif args.operation == 'write':
        # 엑셀 쓰기
        if args.type is None and args.macroname is None:
            wb = xl.load_file(args.filename)
            ws = xl.load_worksheet(wb=wb, sheetname=args.sheetname)
            xl.set_value(wb=wb, ws=ws, cell_range=args.range,
                         filename=args.filename, variable=lst, v=args.vertical)
        # 매크로 실행
        elif args.type is None and args.macroname is not None:
            xl.do_macro(args.filename, args.macroname)
        # 엑셀 함수 사용
        else:
            f = xl.load_csv_file(args.filename)
            data = xl.get_csv_value(file=f)
            data = xl.do_function(data=data, ftype=args.type,
                                  cell_range=args.range)
            xl.to_csv(data=data, filename=args.filename)


if __name__ == "__main__":
    main()
